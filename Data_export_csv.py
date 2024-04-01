import numpy as np
import pandas as pd
import os
from openpyxl import Workbook, load_workbook

#Se importa el archivo CSV UTF-8
df = pd.read_csv("/data/Reporte.csv")
df['FECHA DE RESOLUCIÓN'] = pd.to_datetime(df['FECHA DE RESOLUCIÓN'])

df['day'] = df['FECHA DE RESOLUCIÓN'].dt.day
df['month'] = df['FECHA DE RESOLUCIÓN'].dt.month
df['day'] = df['day'].fillna(0)
df['month']= df['month'].fillna(0)
df['day'] = df['day'].astype(int)
df['month'] = df['month'].astype(int)
d1 = df
d2 = df
d1 = d1.loc[(d1['ESTADO DE LA INCIDENCIA'] == 'Asignado') |
            (d1['ESTADO DE LA INCIDENCIA'] == 'En curso') |
            (d1['ESTADO DE LA INCIDENCIA'] == 'Pendiente')]
d1 = d1.drop('day', axis = 1)
d1 = d1.drop('month', axis = 1)
d1ope = d1.loc[(d1['GRUPO ASIGNADO'] == 'OPE Continuidad') ]
d1teccrm = d1.loc[(d1['GRUPO ASIGNADO'] == 'BACK OFFICE VENTAS') |
                  (d1['GRUPO ASIGNADO'] == 'BACK OFFICE POSTVENTAS') |
                  (d1['GRUPO ASIGNADO'] == 'BACK OFFICE APPS CLIENTES') ]
d1tecboss = d1.loc[(d1['GRUPO ASIGNADO'] == 'BACK OFFICE FACTURACIÓN-CHARGING') |
                  (d1['GRUPO ASIGNADO'] == 'BACK OFFICE PROVISIÓN-MEDIACIÓN') |
                  (d1['GRUPO ASIGNADO'] == 'BACK OFFICE CBIO') ]
d1tecint = d1.loc[(d1['GRUPO ASIGNADO'] == 'BACK OFFICE INTEGRACIÓN') ]
d1soporte = d1.loc[(d1['GRUPO ASIGNADO'] == 'SOPORTE CONFIGURACIONES') ]

columnd1ope = pd.DataFrame([d1ope.columns.tolist()], columns=d1ope.columns)
d1ope = pd.concat([columnd1ope, d1ope])
d1ope = d1ope.reset_index(drop=True)

columnd1teccrm = pd.DataFrame([d1teccrm.columns.tolist()], columns=d1teccrm.columns)
d1teccrm = pd.concat([columnd1teccrm, d1teccrm])
d1teccrm = d1teccrm.reset_index(drop=True)

column_d1tecboss = pd.DataFrame([d1tecboss.columns.tolist()], columns=d1tecboss.columns)
d1tecboss = pd.concat([column_d1tecboss, d1tecboss])
d1tecboss = d1tecboss.reset_index(drop=True)

column_d1tecint = pd.DataFrame([d1tecint.columns.tolist()], columns=d1tecint.columns)
d1tecint = pd.concat([column_d1tecint, d1tecint])
d1tecint = d1tecint.reset_index(drop=True)

column_d1soporte = pd.DataFrame([d1soporte.columns.tolist()], columns=d1soporte.columns)
d1soporte = pd.concat([column_d1soporte, d1soporte])
d1soporte = d1soporte.reset_index(drop=True)


d2 = d2.loc[(d2['ESTADO DE LA INCIDENCIA'] == 'Cerrado') |
            (d2['ESTADO DE LA INCIDENCIA'] == 'Resuelto') ]
d2 = d2.loc[(d2['month'] == 1) ]
d2 = d2.drop('day', axis = 1)
d2 = d2.drop('month', axis = 1)
d2ope = d2.loc[(d2['GRUPO ASIGNADO'] == 'OPE Continuidad')]
d2teccrm = d2.loc[(d2['GRUPO ASIGNADO'] == 'BACK OFFICE VENTAS') |
                  (d2['GRUPO ASIGNADO'] == 'BACK OFFICE POSTVENTAS') |
                  (d2['GRUPO ASIGNADO'] == 'BACK OFFICE APPS CLIENTES')]
d2tecboss = d2.loc[(d2['GRUPO ASIGNADO'] == 'BACK OFFICE FACTURACIÓN-CHARGING') |
                   (d2['GRUPO ASIGNADO'] == 'BACK OFFICE PROVISIÓN-MEDIACIÓN') |
                   (d2['GRUPO ASIGNADO'] == 'BACK OFFICE CBIO')]
d2tecint = d2.loc[(d2['GRUPO ASIGNADO'] == 'BACK OFFICE INTEGRACIÓN')]

column_d2ope = pd.DataFrame([d2ope.columns.tolist()], columns=d2ope.columns)
d2ope = pd.concat([column_d2ope, d2ope])
d2ope = d2ope.reset_index(drop=True)

column_d2teccrm = pd.DataFrame([d2teccrm.columns.tolist()], columns=d2teccrm.columns)
d2teccrm = pd.concat([column_d2teccrm, d2teccrm])
d2teccrm = d2teccrm.reset_index(drop=True)

column_d2tecboss = pd.DataFrame([d2tecboss.columns.tolist()], columns=d2tecboss.columns)
d2tecboss = pd.concat([column_d2tecboss, d2tecboss])
d2tecboss = d2tecboss.reset_index(drop=True)

column_d2tecint = pd.DataFrame([d2tecint.columns.tolist()], columns=d2tecint.columns)
d2tecint = pd.concat([column_d2tecint, d2tecint])
d2tecint = d2tecint.reset_index(drop=True)

d1ope = d1ope.values.tolist()
d1teccrm = d1teccrm.values.tolist()
d1tecboss = d1tecboss.values.tolist()
d1tecint = d1tecint.values.tolist()
d1soporte = d1soporte.values.tolist()

d2ope = d2ope.values.tolist()
d2teccrm = d2teccrm.values.tolist()
d2tecboss = d2tecboss.values.tolist()
d2tecint = d2tecint.values.tolist()

#se crea el archivo tec-crm
wb_teccrm = Workbook()
wb_teccrm1 = wb_teccrm.create_sheet("TEC-CRM", 0)
wb_teccrm2 = wb_teccrm.create_sheet("RESUELTOS TEC-CRM", 1)
wb_teccrm3 = wb_teccrm.create_sheet("OPE CONTINUIDAD", 2)
wb_teccrm4 = wb_teccrm.create_sheet("RESUELTOS OPE CONTINUIDAD", 3)

if len(d1teccrm) != 0:
    for r in d1teccrm:
        wb_teccrm1.append(r)
else:
    del wb_teccrm['TEC-CRM']

if len(d2teccrm) != 0:
    for r in d2teccrm:
        wb_teccrm2.append(r)
else:
    del wb_teccrm['RESUELTOS TEC-CRM']

if len(d1ope[1:]) != 0:
    for r in d1ope:
        wb_teccrm3.append(r)
else:
    del wb_teccrm['OPE CONTINUIDAD']

if len(d2ope[1:]) != 0:
    for r in d2ope:
        wb_teccrm4.append(r)
else:
    del wb_teccrm['RESUELTOS OPE CONTINUIDAD']

del wb_teccrm['Sheet']

wb_teccrm.save("BO_TEC_CRM.xlsx")

#se crear el archivo tec-boss
wb_tecboss = Workbook()
wb_tecboss1 = wb_tecboss.create_sheet("TEC-BOSS", 0)
wb_tecboss2 = wb_tecboss.create_sheet("RESUELTOS TEC-BOSS", 1)
wb_tecboss3 = wb_tecboss.create_sheet("OPE CONTINUIDAD", 2)
wb_tecboss4 = wb_tecboss.create_sheet("RESUELTOS OPE CONTINUIDAD", 3)

if len(d1tecboss) != 0:
    for r in d1tecboss:
        wb_tecboss1.append(r)
else:
    del wb_tecboss['TEC-BOSS']

if len(d2tecboss) != 0:
    for r in d2tecboss:
        wb_tecboss2.append(r)
else:
    del wb_tecboss['RESUELTOS TEC-BOSS']

if len(d1ope[1:]) != 0:
    for r in d1ope:
        wb_tecboss3.append(r)
else:
    del wb_tecboss['OPE CONTINUIDAD']

if len(d2ope[1:]) != 0:
    for r in d2ope:
        wb_tecboss4.append(r)
else:
    del wb_tecboss['RESUELTOS OPE CONTINUIDAD']

del wb_tecboss['Sheet']

wb_tecboss.save("BO_TEC_BOSS.xlsx")

#se crea el archivo BO INTEGRACION
wb_tecint = Workbook()
wb_tecint1 = wb_tecint.create_sheet("BO INTEGRACION", 0)
wb_tecint2 = wb_tecint.create_sheet("RESUELTOS BO INTEGRACION", 1)
wb_tecint3 = wb_tecint.create_sheet("OPE CONTINUIDAD", 2)
wb_tecint4 = wb_tecint.create_sheet("RESUELTOS OPE CONTINUIDAD", 3)

if len(d1tecint) != 0:
    for r in d1tecint:
        wb_tecint1.append(r)
else:
    del wb_tecint['BO INTEGRACION']

if len(d2tecint) != 0:
    for r in d2tecint:
        wb_tecint2.append(r)
else:
    del wb_tecint['RESUELTOS BO INTEGRACION']

if len(d1ope[1:]) != 0:
    for r in d1ope:
        wb_tecint3.append(r)
else:
    del wb_tecint['OPE CONTINUIDAD']

if len(d2ope[1:]) != 0:
    for r in d2ope:
        wb_tecint4.append(r)
else:
    del wb_tecint['RESUELTOS OPE CONTINUIDAD']

del wb_tecint['Sheet']

wb_tecint.save("BO_TEC_INT.xlsx")

#se crea el archivo SOPORTE CONFIGURACIONES
wb_soporte = Workbook()
wb_soporte1 = wb_soporte.create_sheet("SOPORTE CONFIGURACIONES", 0)

for r in d1soporte:
    wb_soporte1.append(r)

wb_soporte.save("BO_TEC_SOPORTE.xlsx")